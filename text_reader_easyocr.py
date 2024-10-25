from openpyxl import Workbook
import cv2
import easyocr

# Initialize the EasyOCR reader
reader = easyocr.Reader(['en'])  # Add more languages if needed
wb = Workbook()

def readWebcamPic(frame):
    # Perform OCR using EasyOCR
    result = reader.readtext(frame)
    # Concatenate the detected text
    text = "\n".join([detection[1] for detection in result])
    if len(text) > 3:
        return text
    return ""

names = ['courses1.jpg', 'courses2.jpg', 'courses3.jpg']

pic_index = 0
row_index = 0
make_excel = True

while pic_index < len(names):
    cap = cv2.VideoCapture(names[pic_index])
    ret, frame = cap.read()
    if not ret:
        print(f"Failed to read image: {names[pic_index]}")
        pic_index += 1
        continue

    frame = cv2.resize(frame, None, fx=0.5, fy=0.5, interpolation=cv2.INTER_AREA)

    try:
        text = readWebcamPic(frame)
        with open('out.txt', 'w') as f:
            print(text, file=f)

        if make_excel:
            sentences = text.split("\n")
            ws = wb.active
            for index in range(len(sentences)):
                # Column A: Store the full text of each sentence
                ws['A' + str(1 + row_index)] = sentences[index]

                # Check if the sentence ends with a price (e.g., "12.34" or "9.99")
                if sentences[index][-2:].isdigit() and sentences[index][-4].isdigit():
                    # Check if it's a price with 3 digits before decimal (e.g., "123.45")
                    if sentences[index][-5:].isdigit():
                        # Column B: Store the item name (everything except the last 11 characters)
                        ws['B' + str(1 + row_index)] = sentences[index][:-11]
                        # Column C: Store the price as a float (e.g., 123.45)
                        ws['C' + str(1 + row_index)] = float(sentences[index][-5:-3]) + float(sentences[index][-2:]) / 100
                    else:
                        # Column B: Store the item name (everything except the last 11 characters)
                        ws['B' + str(1 + row_index)] = sentences[index][:-11]
                        # Column C: Store the price as a float (e.g., 12.34)
                        ws['C' + str(1 + row_index)] = float(sentences[index][-4]) + float(sentences[index][-2:]) / 100
                else:
                    # If no price is found, store the full text in column B and 'No price' in column C
                    ws['B' + str(1 + row_index)] = sentences[index]
                    ws['C' + str(1 + row_index)] = 'No price'
                row_index += 1

            # Add a separator between different images
            if pic_index < len(names) - 1:
                row_index += 1
                # Add a label for the next image
                ws['A' + str(1 + row_index)] = 'Photo numero ' + str(pic_index + 2)
                row_index += 3
                pic_index += 1
            else:
                pic_index += 1
        else:
            pic_index += 1
    except Exception as e:
        print(f"An error occurred: {e}")

if make_excel:
    wb.save('sample.xlsx')

cap.release()
cv2.destroyAllWindows()