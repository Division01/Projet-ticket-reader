from openpyxl import Workbook
import cv2
import pytesseract

# Set the path to the Tesseract executable
pytesseract.pytesseract.tesseract_cmd = r'E:\Code\Tesseract (OCR)\\tesseract.exe'
wb = Workbook()

def readWebcamPic(frame):
    text = pytesseract.image_to_string(frame)
    if len(text) > 3:
        return text
    return ""

names = ['courses1.jpg', 'courses2.jpg', 'courses3.jpg']

pic_index = 0
row_index = 0
make_excel = True

while pic_index < len(names):
    cap = cv2.VideoCapture(names[pic_index])
    ret, frame = cap.read()
    if not ret:
        print(f"Failed to read image: {names[pic_index]}")
        pic_index += 1
        continue

    frame = cv2.resize(frame, None, fx=0.5, fy=0.5, interpolation=cv2.INTER_AREA)

    try:
        text = readWebcamPic(frame)
        with open('out.txt', 'w') as f:
            print(text, file=f)

        if make_excel:
            sentences = text.split("\n")
            ws = wb.active
            for index in range(len(sentences)):
                ws['A' + str(1 + row_index)] = sentences[index]

                if sentences[index][-2:].isdigit() and sentences[index][-4].isdigit():
                    if sentences[index][-5:].isdigit():
                        ws['B' + str(1 + row_index)] = sentences[index][:-11]
                        ws['C' + str(1 + row_index)] = float(sentences[index][-5:-3]) + float(sentences[index][-2:]) / 100
                    else:
                        ws['B' + str(1 + row_index)] = sentences[index][:-11]
                        ws['C' + str(1 + row_index)] = float(sentences[index][-4]) + float(sentences[index][-2:]) / 100
                else:
                    ws['B' + str(1 + row_index)] = sentences[index]
                    ws['C' + str(1 + row_index)] = 'No price'
                row_index += 1

            if pic_index < len(names) - 1:
                row_index += 1
                ws['A' + str(1 + row_index)] = 'Photo numero ' + str(pic_index + 2)
                row_index += 3
                pic_index += 1
            else:
                pic_index += 1
        else:
            pic_index += 1
    except Exception as e:
        print(f"An error occurred while processing {names[pic_index]}: {e}")
        pic_index += 1

if make_excel:
    wb.save('sample.xlsx')

cap.release()
cv2.destroyAllWindows()